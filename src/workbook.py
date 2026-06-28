"""Assemble outputs/results.xlsx from the computed tables and figure data.

Design choice for auditability: the classification metrics (F1, precision,
recall, FPR, FNR) are written as Excel FORMULAS over the confusion-matrix counts
(TP/FP/FN/TN), not as hardcoded numbers. A reviewer can see exactly how each
metric follows from the counts, and the sheet recomputes if the counts change.
Counts come from the per-instance predictions in data/ via src/metrics.py.

Measured quantities that are not derivable from counts (latency, memory,
parameter counts, AUC, SHAP magnitudes, static-tool points) are written as
values and labelled as direct inputs.
"""
from __future__ import annotations
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from metrics import confusion_at_threshold, select_threshold, roc_curve

DATA = os.path.join(os.path.dirname(__file__), "..", "data")
OUT = os.path.join(os.path.dirname(__file__), "..", "outputs")

HDR = Font(bold=True, color="FFFFFF", name="Arial", size=10)
HDR_FILL = PatternFill("solid", start_color="33445B")
TITLE = Font(bold=True, size=12, name="Arial")
NOTE = Font(italic=True, color="B00000", name="Arial", size=10)
BASE = Font(name="Arial", size=10)
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def _style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")


def _counts(pred_csv, key):
    df = pd.read_csv(pred_csv)
    out = []
    for k, g in df.groupby(["task", key] if key != "task" else ["task"]):
        thr = select_threshold(g["y_true"], g["y_score"], "default")
        c = confusion_at_threshold(g["y_true"], g["y_score"], thr)
        rec = {"task": k[0]} if isinstance(k, tuple) else {"task": k}
        if key != "task" and isinstance(k, tuple):
            rec[key] = k[1]
        rec.update(c)
        out.append(rec)
    return pd.DataFrame(out)


def _metric_block(ws, start_row, label_cols, counts_df):
    """Write a counts table with formula-derived metric columns. Returns next row."""
    headers = label_cols + ["TP", "FP", "FN", "TN",
                            "Precision %", "Recall %", "F1 %", "FPR %", "FNR %"]
    r = start_row
    for j, h in enumerate(headers, 1):
        ws.cell(row=r, column=j, value=h)
    _style_header(ws, r, len(headers))
    r += 1
    nlab = len(label_cols)
    for _, row in counts_df.iterrows():
        for j, lc in enumerate(label_cols, 1):
            ws.cell(row=r, column=j, value=row[lc]).font = BASE
        tpc, fpc, fnc, tnc = (nlab + 1, nlab + 2, nlab + 3, nlab + 4)
        ws.cell(row=r, column=tpc, value=int(row["tp"])).font = BASE
        ws.cell(row=r, column=fpc, value=int(row["fp"])).font = BASE
        ws.cell(row=r, column=fnc, value=int(row["fn"])).font = BASE
        ws.cell(row=r, column=tnc, value=int(row["tn"])).font = BASE
        TP = ws.cell(row=r, column=tpc).coordinate
        FP = ws.cell(row=r, column=fpc).coordinate
        FN = ws.cell(row=r, column=fnc).coordinate
        TN = ws.cell(row=r, column=tnc).coordinate
        ws.cell(row=r, column=nlab + 5,
                value=f"=IF(({TP}+{FP})=0,0,100*{TP}/({TP}+{FP}))").font = BASE
        ws.cell(row=r, column=nlab + 6,
                value=f"=IF(({TP}+{FN})=0,0,100*{TP}/({TP}+{FN}))").font = BASE
        ws.cell(row=r, column=nlab + 7,
                value=f"=IF((2*{TP}+{FP}+{FN})=0,0,100*2*{TP}/(2*{TP}+{FP}+{FN}))").font = BASE
        ws.cell(row=r, column=nlab + 8,
                value=f"=IF(({FP}+{TN})=0,0,100*{FP}/({FP}+{TN}))").font = BASE
        ws.cell(row=r, column=nlab + 9,
                value=f"=IF(({FN}+{TP})=0,0,100*{FN}/({FN}+{TP}))").font = BASE
        r += 1
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=start_row, column=col).column_letter].width = 14
    return r + 1


def build():
    os.makedirs(OUT, exist_ok=True)
    wb = Workbook()

    # --- README sheet ---
    ws = wb.active; ws.title = "README"
    ws["A1"] = "TEFIE-Secure -- results workbook"; ws["A1"].font = TITLE
    msg = [
        "Classification metrics (Accuracy, Ablation, RealWorld sheets) are Excel FORMULAS",
        "over the TP/FP/FN/TN counts, so each metric is auditable in-cell.",
        "Counts are produced by src/metrics.py from the per-instance predictions in data/.",
        "",
        "The data/ shipped with this artifact is SYNTHETIC smoke-test data and DOES NOT",
        "reproduce the values in the manuscript. Replace data/ with your real experimental",
        "outputs (schema in data/SCHEMA.md) and re-run reproduce.py to obtain the paper's",
        "tables and figures.",
    ]
    for i, line in enumerate(msg, start=3):
        ws.cell(row=i, column=1, value=line).font = NOTE if "SYNTHETIC" in line else BASE
    ws.column_dimensions["A"].width = 95

    # --- Accuracy ---
    ws = wb.create_sheet("Accuracy")
    ws["A1"] = "Detection accuracy (default threshold; metrics are formulas of the counts)"
    ws["A1"].font = TITLE
    _metric_block(ws, 3, ["task", "model"],
                  _counts(os.path.join(DATA, "predictions_classification.csv"), "model"))

    # --- Ablation ---
    ws = wb.create_sheet("Ablation")
    ws["A1"] = "Ablation (F1 derived in-cell from counts)"; ws["A1"].font = TITLE
    _metric_block(ws, 3, ["task", "variant"],
                  _counts(os.path.join(DATA, "predictions_ablation.csv"), "variant"))

    # --- Runtime (measured inputs) ---
    ws = wb.create_sheet("Runtime")
    ws["A1"] = "Runtime / resource profile (direct measurements)"; ws["A1"].font = TITLE
    rt = pd.read_csv(os.path.join(DATA, "runtime.csv"))
    for j, h in enumerate(rt.columns, 1):
        ws.cell(row=3, column=j, value=h)
    _style_header(ws, 3, len(rt.columns))
    for i, (_, row) in enumerate(rt.iterrows(), start=4):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v).font = BASE
    for col in range(1, len(rt.columns) + 1):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = 16

    # --- Real-world (RQ5) ---
    ws = wb.create_sheet("RealWorld")
    ws["A1"] = "RQ5 post-2023 mainnet (high-recall threshold)"; ws["A1"].font = TITLE
    rw = pd.read_csv(os.path.join(DATA, "realworld.csv"))
    thr = select_threshold(rw["y_true"], rw["y_score"], "high_recall", target=0.02)
    rw = rw.assign(flagged=(rw["y_score"] >= thr).astype(int))
    disclosed = rw[rw["disclosed"] == 1]
    bt = (disclosed.groupby("vuln_type")
          .agg(disclosed=("y_true", "size"), flagged=("flagged", "sum")).reset_index())
    ws.cell(row=3, column=1, value="vuln_type"); ws.cell(row=3, column=2, value="disclosed")
    ws.cell(row=3, column=3, value="flagged"); ws.cell(row=3, column=4, value="recall %")
    _style_header(ws, 3, 4)
    r = 4
    for _, row in bt.iterrows():
        ws.cell(row=r, column=1, value=row["vuln_type"]).font = BASE
        ws.cell(row=r, column=2, value=int(row["disclosed"])).font = BASE
        ws.cell(row=r, column=3, value=int(row["flagged"])).font = BASE
        B = ws.cell(row=r, column=2).coordinate; C = ws.cell(row=r, column=3).coordinate
        ws.cell(row=r, column=4, value=f"=IF({B}=0,0,100*{C}/{B})").font = BASE
        r += 1
    tp = int(disclosed["flagged"].sum())
    fp_weak = int(rw[(rw["disclosed"] == 0) & (rw["flagged"] == 1)].shape[0])
    r += 1
    ws.cell(row=r, column=1, value="disclosed_flagged (TP)").font = BASE
    ws.cell(row=r, column=2, value=tp).font = BASE
    ws.cell(row=r + 1, column=1, value="weak_negative_flags (FP)").font = BASE
    ws.cell(row=r + 1, column=2, value=fp_weak).font = BASE
    ws.cell(row=r + 2, column=1, value="weak_negative_precision %").font = Font(bold=True, name="Arial")
    tpc = ws.cell(row=r, column=2).coordinate; fpc = ws.cell(row=r + 1, column=2).coordinate
    ws.cell(row=r + 2, column=2, value=f"=IF(({tpc}+{fpc})=0,0,100*{tpc}/({tpc}+{fpc}))").font = Font(bold=True, name="Arial")
    for col in "ABCD":
        ws.column_dimensions[col].width = 24

    # --- Graph data: ROC AUC + static points ---
    ws = wb.create_sheet("Graph_ROC")
    ws["A1"] = "ROC data (AUC computed by metrics.roc_curve; static points are inputs)"
    ws["A1"].font = TITLE
    ws.cell(row=3, column=1, value="task"); ws.cell(row=3, column=2, value="model")
    ws.cell(row=3, column=3, value="AUC"); _style_header(ws, 3, 3)
    preds = pd.read_csv(os.path.join(DATA, "predictions_classification.csv"))
    r = 4
    for task in ["reentrancy", "timestamp", "loops"]:
        for model, g in preds[preds["task"] == task].groupby("model"):
            _, _, auc = roc_curve(g["y_true"], g["y_score"])
            ws.cell(row=r, column=1, value=task).font = BASE
            ws.cell(row=r, column=2, value=model).font = BASE
            ws.cell(row=r, column=3, value=round(auc, 3)).font = BASE
            r += 1
    sp = pd.read_csv(os.path.join(DATA, "static_points.csv"))
    r += 1; ws.cell(row=r, column=1, value="Static-tool points (FPR, TPR)").font = Font(bold=True, name="Arial"); r += 1
    ws.cell(row=r, column=1, value="task"); ws.cell(row=r, column=2, value="tool")
    ws.cell(row=r, column=3, value="FPR"); ws.cell(row=r, column=4, value="TPR")
    _style_header(ws, r, 4); r += 1
    for _, row in sp.iterrows():
        ws.cell(row=r, column=1, value=row["task"]).font = BASE
        ws.cell(row=r, column=2, value=row["tool"]).font = BASE
        ws.cell(row=r, column=3, value=row["fpr"]).font = BASE
        ws.cell(row=r, column=4, value=row["tpr"]).font = BASE
        r += 1
    for col in "ABCD":
        ws.column_dimensions[col].width = 16

    # --- Graph data: SHAP ---
    ws = wb.create_sheet("Graph_SHAP")
    ws["A1"] = "Feature-importance data (mean |SHAP| inputs)"; ws["A1"].font = TITLE
    sh = pd.read_csv(os.path.join(DATA, "shap.csv"))
    for j, h in enumerate(sh.columns, 1):
        ws.cell(row=3, column=j, value=h)
    _style_header(ws, 3, len(sh.columns))
    for i, (_, row) in enumerate(sh.iterrows(), start=4):
        for j, v in enumerate(row, 1):
            ws.cell(row=i, column=j, value=v).font = BASE
    for col in "ABC":
        ws.column_dimensions[col].width = 18

    path = os.path.join(OUT, "results.xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    print("workbook written to", build())
